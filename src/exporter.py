"""Multi-format exporter for the Arabic Medical Glossary."""

import csv
import json
import logging
import sqlite3
from datetime import datetime, timezone
from io import StringIO
from pathlib import Path
from typing import Any, Optional

from .config import (
    API_CONFIG,
    BACKUP_DIR,
    DB_PATH,
    HIGH_CONFIDENCE_THRESHOLD,
    PROJECT_ROOT,
)

logger = logging.getLogger(__name__)


class MultiExporter:
    """Export glossary data to CSV, JSON, JSONL, TSV, SQLite, Excel, HuggingFace, FastText, TMX."""

    def __init__(self, db_manager: Any = None) -> None:
        self.db = db_manager
        self._export_log: list[dict[str, Any]] = []

    def _get_terms(self) -> list[dict]:
        if self.db:
            return self.db.search(limit=1_000_000)
        return []

    def _log(self, fmt: str, path: str, count: int) -> None:
        entry = {
            "format": fmt,
            "path": str(path),
            "count": count,
            "timestamp": datetime.now(timezone.utc).isoformat(),
        }
        self._export_log.append(entry)
        logger.info("Exported %d terms to %s (%s)", count, path, fmt)

    # ------------------------------------------------------------------
    # CSV (UTF-8 BOM)
    # ------------------------------------------------------------------
    def export_to_csv(
        self, output_path: str | Path, terms: Optional[list[dict]] = None
    ) -> Path:
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        data = terms or self._get_terms()
        if not data:
            logger.warning("No terms to export")
            return output_path

        fieldnames = list(data[0].keys())
        with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(data)
        self._log("csv", output_path, len(data))
        return output_path

    # ------------------------------------------------------------------
    # JSON
    # ------------------------------------------------------------------
    def export_to_json(
        self, output_path: str | Path, terms: Optional[list[dict]] = None
    ) -> Path:
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        data = terms or self._get_terms()
        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        self._log("json", output_path, len(data))
        return output_path

    # ------------------------------------------------------------------
    # JSONL
    # ------------------------------------------------------------------
    def export_to_jsonl(
        self, output_path: str | Path, terms: Optional[list[dict]] = None
    ) -> Path:
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        data = terms or self._get_terms()
        with open(output_path, "w", encoding="utf-8") as f:
            for row in data:
                f.write(json.dumps(row, ensure_ascii=False) + "\n")
        self._log("jsonl", output_path, len(data))
        return output_path

    # ------------------------------------------------------------------
    # TSV
    # ------------------------------------------------------------------
    def export_to_tsv(
        self, output_path: str | Path, terms: Optional[list[dict]] = None
    ) -> Path:
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        data = terms or self._get_terms()
        if not data:
            return output_path
        fieldnames = list(data[0].keys())
        with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames, delimiter="\t")
            writer.writeheader()
            writer.writerows(data)
        self._log("tsv", output_path, len(data))
        return output_path

    # ------------------------------------------------------------------
    # Standalone SQLite with FTS5
    # ------------------------------------------------------------------
    def export_to_sqlite(
        self, output_path: str | Path, terms: Optional[list[dict]] = None
    ) -> Path:
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        data = terms or self._get_terms()
        if output_path.exists():
            output_path.unlink()

        conn = sqlite3.connect(str(output_path))
        conn.execute("PRAGMA journal_mode=WAL")
        if data:
            fieldnames = list(data[0].keys())
            cols = ", ".join(f'"{c}" TEXT' for c in fieldnames)
            conn.execute(f"CREATE TABLE terms ({cols})")
            placeholders = ", ".join("?" for _ in fieldnames)
            rows = [tuple(str(r.get(c, "")) for c in fieldnames) for r in data]
            conn.executemany(f"INSERT INTO terms VALUES ({placeholders})", rows)

            # FTS5
            fts_cols = ", ".join(
                f'"{c}"' for c in fieldnames
                if c in ("english", "arabic", "category", "source", "section")
            )
            if fts_cols:
                conn.execute(
                    f"CREATE VIRTUAL TABLE terms_fts USING fts5({fts_cols}, content='terms', content_rowid='rowid')"
                )
                idx_cols = [c for c in fieldnames if c in ("english", "arabic", "category", "source", "section")]
                sel_cols = ", ".join(f'"{c}"' for c in idx_cols)
                conn.execute(f"INSERT INTO terms_fts({sel_cols}) SELECT {sel_cols} FROM terms")

        conn.commit()
        conn.close()
        self._log("sqlite", output_path, len(data))
        return output_path

    # ------------------------------------------------------------------
    # Excel (openpyxl with RTL)
    # ------------------------------------------------------------------
    def export_to_excel(
        self, output_path: str | Path, terms: Optional[list[dict]] = None
    ) -> Path:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment
        except ImportError:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")

        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        data = terms or self._get_terms()

        wb = Workbook()
        ws = wb.active
        ws.title = "Glossary"

        if data:
            headers = list(data[0].keys())
            ws.append(headers)

            # Set RTL and header style
            rtl_align = Alignment(reading_order=2, wrap_text=True)
            for cell in ws[1]:
                cell.alignment = rtl_align

            for row_data in data:
                ws.append([row_data.get(h, "") for h in headers])

            # Auto-adjust column widths
            for col_idx, header in enumerate(headers, 1):
                max_len = len(str(header))
                for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        if cell.value:
                            max_len = max(max_len, min(len(str(cell.value)), 60))
                ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max_len + 4

        wb.save(str(output_path))
        self._log("xlsx", output_path, len(data))
        return output_path

    # ------------------------------------------------------------------
    # HuggingFace datasets
    # ------------------------------------------------------------------
    def export_to_huggingface(
        self, output_dir: str | Path, terms: Optional[list[dict]] = None,
        test_size: float = 0.1
    ) -> Path:
        try:
            from datasets import Dataset, DatasetDict
        except ImportError:
            raise ImportError("datasets library is required for HuggingFace export")

        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)
        data = terms or self._get_terms()
        if not data:
            return output_dir

        ds = Dataset.from_list(data)
        split = ds.train_test_split(test_size=test_size, seed=42)
        ds_dict: DatasetDict = DatasetDict({"train": split["train"], "test": split["test"]})
        ds_dict.save_to_disk(str(output_dir))
        self._log("huggingface", output_dir, len(data))
        return output_dir

    # ------------------------------------------------------------------
    # FastText
    # ------------------------------------------------------------------
    def export_to_fasttext(
        self, output_path: str | Path, terms: Optional[list[dict]] = None
    ) -> Path:
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        data = terms or self._get_terms()
        with open(output_path, "w", encoding="utf-8") as f:
            for t in data:
                en = str(t.get("english", "")).replace("\n", " ")
                ar = str(t.get("arabic", "")).replace("\n", " ")
                src = str(t.get("source", "unknown"))
                f.write(f"__label__{src} {en}\t{ar}\n")
        self._log("fasttext", output_path, len(data))
        return output_path

    # ------------------------------------------------------------------
    # TMX (Translation Memory XML)
    # ------------------------------------------------------------------
    def export_to_tmx(
        self, output_path: str | Path, terms: Optional[list[dict]] = None
    ) -> Path:
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        data = terms or self._get_terms()

        lines: list[str] = [
            '<?xml version="1.0" encoding="UTF-8"?>',
            '<!DOCTYPE tmx SYSTEM "tmx14.dtd">',
            '<tmx version="1.4">',
            '  <header creationtool="arabic-medical-glossary" creationtoolversion="1.0"',
            '          datatype="plaintext" segtype="sentence" adminlang="en-US"',
            '          srclang="en" o-tmf="arabic-medical-glossary">',
            '  </header>',
            '  <body>',
        ]
        for t in data:
            en = str(t.get("english", "")).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
            ar = str(t.get("arabic", "")).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
            lines.append("    <tu>")
            lines.append(f'      <tuv xml:lang="en"><seg>{en}</seg></tuv>')
            lines.append(f'      <tuv xml:lang="ar"><seg>{ar}</seg></tuv>')
            lines.append("    </tu>")
        lines.extend(["  </body>", "</tmx>"])

        with open(output_path, "w", encoding="utf-8") as f:
            f.write("\n".join(lines) + "\n")
        self._log("tmx", output_path, len(data))
        return output_path

    # ------------------------------------------------------------------
    # Export all formats at once
    # ------------------------------------------------------------------
    def export_all_formats(
        self, output_dir: str | Path, terms: Optional[list[dict]] = None
    ) -> dict[str, Path]:
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)
        results: dict[str, Path] = {}
        data = terms or self._get_terms()

        results["csv"] = self.export_to_csv(output_dir / "glossary.csv", data)
        results["json"] = self.export_to_json(output_dir / "glossary.json", data)
        results["jsonl"] = self.export_to_jsonl(output_dir / "glossary.jsonl", data)
        results["tsv"] = self.export_to_tsv(output_dir / "glossary.tsv", data)
        results["sqlite"] = self.export_to_sqlite(output_dir / "glossary_export.db", data)
        results["tmx"] = self.export_to_tmx(output_dir / "glossary.tmx", data)
        results["fasttext"] = self.export_to_fasttext(output_dir / "glossary_fasttext.txt", data)
        try:
            results["xlsx"] = self.export_to_excel(output_dir / "glossary.xlsx", data)
        except ImportError:
            logger.warning("openpyxl not available, skipping Excel export")
        return results

    def get_export_log(self) -> list[dict[str, Any]]:
        return list(self._export_log)